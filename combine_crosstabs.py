"""
Trim each crosstab .xlsx in a survey folder down to just its 'Percentages' sheet
(renamed to the file's name), then combine all of them into one workbook —
one sheet per crosstab.
"""

import copy
from pathlib import Path

import openpyxl

SHEETS_TO_DROP = {"Summary", "Counts"}
KEEP_SHEET = "Percentages"


def trim_file(path: Path) -> str:
    wb = openpyxl.load_workbook(path)
    if KEEP_SHEET not in wb.sheetnames:
        raise ValueError(f"{path.name}: no '{KEEP_SHEET}' sheet found (has {wb.sheetnames})")
    for name in list(wb.sheetnames):
        if name in SHEETS_TO_DROP:
            del wb[name]
    new_name = path.stem[:31]  # Excel sheet name length limit
    wb[KEEP_SHEET].title = new_name
    wb.save(path)
    return new_name


def copy_sheet(src_ws, dst_wb, sheet_name):
    dst_ws = dst_wb.create_sheet(title=sheet_name)

    for row in src_ws.iter_rows():
        for cell in row:
            new_cell = dst_ws.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                new_cell.font = copy.copy(cell.font)
                new_cell.fill = copy.copy(cell.fill)
                new_cell.border = copy.copy(cell.border)
                new_cell.alignment = copy.copy(cell.alignment)
                new_cell.number_format = cell.number_format
                new_cell.protection = copy.copy(cell.protection)

    for merged_range in src_ws.merged_cells.ranges:
        dst_ws.merge_cells(str(merged_range))

    for col_letter, dim in src_ws.column_dimensions.items():
        dst_dim = dst_ws.column_dimensions[col_letter]
        dst_dim.width = dim.width
        dst_dim.hidden = dim.hidden

    for row_idx, dim in src_ws.row_dimensions.items():
        dst_dim = dst_ws.row_dimensions[row_idx]
        dst_dim.height = dim.height
        dst_dim.hidden = dim.hidden

    dst_ws.sheet_view.rightToLeft = src_ws.sheet_view.rightToLeft
    dst_ws.freeze_panes = src_ws.freeze_panes


def run_combine(survey_folder, output: str = None, progress=None) -> Path:
    """Trim every .xlsx in survey_folder down to its Percentages sheet (renamed to the
    file name), then combine them all into one workbook. Returns the combined file's path."""
    def log(msg):
        if progress:
            progress(msg)

    folder = Path(survey_folder)
    if not folder.is_dir():
        raise NotADirectoryError(f"Not a folder: {folder}")

    output_path = Path(output) if output else folder / f"{folder.name} - combined.xlsx"

    xlsx_files = sorted(
        p for p in folder.glob("*.xlsx")
        if p.resolve() != output_path.resolve() and not p.name.startswith("~$")
    )
    if not xlsx_files:
        raise FileNotFoundError(f"No .xlsx files found in {folder}")

    log(f"Trimming {len(xlsx_files)} file(s) to just the '{KEEP_SHEET}' sheet...")
    trimmed_names = []
    for path in xlsx_files:
        sheet_name = trim_file(path)
        trimmed_names.append(sheet_name)
        log(f"  {path.name} -> sheet '{sheet_name}'")

    log(f"Combining into {output_path.name} ...")
    combined_wb = openpyxl.Workbook()
    combined_wb.remove(combined_wb.active)  # drop the default empty sheet

    for path, sheet_name in zip(xlsx_files, trimmed_names):
        src_wb = openpyxl.load_workbook(path)
        copy_sheet(src_wb[sheet_name], combined_wb, sheet_name)

    combined_wb.save(output_path)
    log(f"Done. {len(trimmed_names)} sheet(s) in {output_path.name}")
    return output_path
